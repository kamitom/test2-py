

import datetime
import random
from openpyxl import Workbook

wb2 = Workbook()

ws = wb2.active

ws['C1'] = random.randint(1, 999)
ws['A1'] = random.randint(2, 888)

# TypeError: unsupported operand type(s) for *: 'Cell' and 'Cell'
ws['A5'] = 100 + 2


# ws.append([1, 2, 3])

# ws['A2'] = datetime.datetime.now()

fname = random.randint(1, 1000)
print("obj type: " + str(type(fname)))
print(fname)

filename = str(fname) + ".xlsx"

wb2.save(filename)

wb2.close()
